"""Quantities export — Excel and PDF generation (same tree as quantities panel)."""

from __future__ import annotations

import io
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.middleware.error_handler import NotFoundException
from app.models.condition import Condition
from app.models.measurement import Measurement
from app.models.project import Project
from app.models.sheet import Sheet


@dataclass
class ExportMeasurementRow:
    label: str
    quantity: str
    unit: str


@dataclass
class ExportSheetBlock:
    sheet_name: str
    subtotal_display: str
    unit: str
    measurements: list[ExportMeasurementRow] = field(default_factory=list)


@dataclass
class ExportConditionBlock:
    condition_name: str
    grand_total_display: str
    unit: str
    sheets: list[ExportSheetBlock] = field(default_factory=list)


@dataclass
class ExportBundle:
    project_name: str
    exported_at: datetime
    condition_count: int
    measurement_count: int
    sheet_count: int
    blocks: list[ExportConditionBlock]
    naive_numeric_sum: float


def _effective_value(m: Measurement) -> float:
    return float(m.override_value) if m.override_value is not None else float(m.measured_value)


def _fmt_simple_value(val: float, c: Condition) -> str:
    if c.measurement_type == "count":
        return str(int(round(val)))
    return f"{val:,.2f}"


def _fmt_quantity(m: Measurement, c: Condition) -> str:
    eff = _effective_value(m)
    mv = float(m.measured_value)
    if c.measurement_type == "count":
        q = str(int(round(eff)))
        if m.override_value is not None:
            return f"{q} ({int(round(mv))})"
        return q
    if c.measurement_type == "area":
        q = f"{eff:,.2f}"
        if m.override_value is not None:
            return f"{q} ({mv:,.2f})"
        return q
    # linear
    q = f"{eff:,.2f}"
    if m.override_value is not None:
        return f"{q} ({mv:,.2f})"
    return q


def _measurement_label(m: Measurement) -> str:
    if m.label and str(m.label).strip():
        return str(m.label).strip()
    return f"Measurement {str(m.id)[:8]}"


def _sheet_label(sh: Sheet) -> str:
    return sh.sheet_name or f"Page {sh.page_number}"


def load_export_bundle_sync(session: Session, org_id: uuid.UUID, project_id: uuid.UUID) -> ExportBundle:
    proj = session.execute(
        select(Project).where(Project.id == project_id, Project.org_id == org_id)
    ).scalar_one_or_none()
    if not proj:
        raise NotFoundException("project", str(project_id))

    conditions = list(
        session.execute(
            select(Condition)
            .where(Condition.project_id == project_id, Condition.org_id == org_id)
            .order_by(Condition.sort_order.asc(), Condition.name.asc())
        ).scalars().all()
    )
    measurements = list(
        session.execute(
            select(Measurement)
            .where(Measurement.project_id == project_id, Measurement.org_id == org_id)
            .order_by(Measurement.created_at.asc())
        ).scalars().all()
    )
    sheets = list(
        session.execute(select(Sheet).where(Sheet.project_id == project_id, Sheet.org_id == org_id)).scalars().all()
    )
    sheet_by_id: dict[uuid.UUID, Sheet] = {s.id: s for s in sheets}

    blocks: list[ExportConditionBlock] = []
    naive_sum = 0.0
    for cond in conditions:
        ms = [m for m in measurements if m.condition_id == cond.id]
        if not ms:
            continue
        sheet_ids = sorted(
            {m.sheet_id for m in ms},
            key=lambda sid: (sheet_by_id[sid].page_number if sid in sheet_by_id else 0, str(sid)),
        )
        sheet_blocks: list[ExportSheetBlock] = []
        cond_total = 0.0
        for sid in sheet_ids:
            sh = sheet_by_id.get(sid)
            if not sh:
                continue
            in_sheet = [m for m in ms if m.sheet_id == sid]
            sub = sum(_effective_value(m) for m in in_sheet)
            cond_total += sub
            rows = [
                ExportMeasurementRow(
                    label=_measurement_label(m),
                    quantity=_fmt_quantity(m, cond),
                    unit=cond.unit,
                )
                for m in in_sheet
            ]
            sub_disp = _fmt_simple_value(sub, cond)
            sheet_blocks.append(
                ExportSheetBlock(
                    sheet_name=_sheet_label(sh),
                    subtotal_display=sub_disp,
                    unit=cond.unit,
                    measurements=rows,
                )
            )

        gstr = _fmt_simple_value(cond_total, cond)
        blocks.append(
            ExportConditionBlock(
                condition_name=cond.name,
                grand_total_display=gstr,
                unit=cond.unit,
                sheets=sheet_blocks,
            )
        )
        naive_sum += cond_total

    unique_sheets = len({m.sheet_id for m in measurements})
    return ExportBundle(
        project_name=proj.name,
        exported_at=datetime.now(timezone.utc),
        condition_count=len(blocks),
        measurement_count=len(measurements),
        sheet_count=unique_sheets,
        blocks=blocks,
        naive_numeric_sum=naive_sum,
    )


def build_xlsx_bytes(bundle: ExportBundle) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Quantities"

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    cond_font = Font(bold=True)
    sub_font = Font(bold=False, italic=True)
    fill_hdr = PatternFill("solid", fgColor="1e293b")
    font_hdr = Font(color="f8fafc", bold=True)

    ws["A1"] = "Contruo — Quantities export"
    ws["A1"].font = title_font
    ws["A2"] = f"Project: {bundle.project_name}"
    ws["B2"] = f"Exported: {bundle.exported_at.strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A3"] = (
        f"Summary: {bundle.condition_count} conditions · {bundle.measurement_count} measurements · "
        f"{bundle.sheet_count} sheets"
    )

    header_row = 5
    ws.cell(row=header_row, column=1, value="Name / Label")
    ws.cell(row=header_row, column=2, value="Quantity")
    ws.cell(row=header_row, column=3, value="Unit")
    for c in range(1, 4):
        cell = ws.cell(row=header_row, column=c)
        cell.font = font_hdr
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row = header_row + 1
    widths = [40.0, 18.0, 10.0]

    for block in bundle.blocks:
        ws.cell(row=row, column=1, value=f"{block.condition_name} — Total: {block.grand_total_display} {block.unit}")
        ws.cell(row=row, column=1).font = cond_font
        ws.row_dimensions[row].outline_level = 0
        row += 1
        for sb in block.sheets:
            ws.cell(
                row=row,
                column=1,
                value=f"  {sb.sheet_name} — Subtotal: {sb.subtotal_display}",
            )
            ws.cell(row=row, column=1).font = sub_font
            ws.row_dimensions[row].outline_level = 1
            row += 1
            for mr in sb.measurements:
                ws.cell(row=row, column=1, value=f"    {mr.label}")
                ws.cell(row=row, column=2, value=mr.quantity)
                ws.cell(row=row, column=3, value=mr.unit)
                ws.row_dimensions[row].outline_level = 2
                for col, val in enumerate([mr.label, mr.quantity, mr.unit], start=1):
                    ln = len(str(val))
                    if col <= len(widths):
                        widths[col - 1] = max(widths[col - 1], min(60, ln + 2))
                row += 1

    ws.cell(row=row, column=1, value="Grand total (numeric sum — verify units)")
    ws.cell(row=row, column=2, value=f"{bundle.naive_numeric_sum:,.4f}")
    ws.cell(row=row, column=1).font = Font(bold=True)

    try:
        ws.sheet_properties.outlinePr.summaryBelow = False  # type: ignore[union-attr]
    except Exception:
        pass
    ws.column_dimensions["A"].width = widths[0]
    ws.column_dimensions["B"].width = widths[1]
    ws.column_dimensions["C"].width = widths[2]

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_pdf_bytes(bundle: ExportBundle) -> bytes:
    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(letter),
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.65 * inch,
        bottomMargin=0.65 * inch,
        title="Quantities export",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="ContruoTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=6,
        textColor=colors.HexColor("#0f172a"),
    )
    meta_style = ParagraphStyle(
        name="Meta",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#475569"),
    )

    story: list[Any] = []
    story.append(Paragraph("Contruo — Quantities export", title_style))
    story.append(Paragraph(f"<b>Project:</b> {bundle.project_name}", meta_style))
    story.append(
        Paragraph(
            f"Exported {bundle.exported_at.strftime('%Y-%m-%d %H:%M UTC')} · "
            f"{bundle.condition_count} conditions · {bundle.measurement_count} measurements · "
            f"{bundle.sheet_count} sheets",
            meta_style,
        )
    )
    story.append(Spacer(1, 0.2 * inch))

    table_data: list[list[Any]] = [["Name / Label", "Quantity", "Unit"]]
    for block in bundle.blocks:
        table_data.append([f"{block.condition_name} — Total: {block.grand_total_display} {block.unit}", "", ""])
        for sb in block.sheets:
            table_data.append([f"  {sb.sheet_name} — Subtotal: {sb.subtotal_display}", "", ""])
            for mr in sb.measurements:
                table_data.append([f"    {mr.label}", mr.quantity, mr.unit])
    table_data.append(
        ["Grand total (numeric sum — verify units)", f"{bundle.naive_numeric_sum:,.4f}", ""]
    )

    tbl = Table(table_data, colWidths=[5.8 * inch, 1.6 * inch, 0.9 * inch], repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (1, 0), (2, -1), "RIGHT"),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(tbl)

    def _draw_header_footer(canvas: Any, doc_: Any) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(doc_.leftMargin, doc_.height + doc_.bottomMargin - 0.35 * inch, bundle.project_name[:80])
        canvas.drawRightString(
            doc_.leftMargin + doc_.width,
            doc_.bottomMargin - 0.25 * inch,
            f"Page {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=_draw_header_footer, onLaterPages=_draw_header_footer)
    return bio.getvalue()


def generate_export_bytes(session: Session, org_id: uuid.UUID, project_id: uuid.UUID, export_format: str) -> tuple[bytes, str, str]:
    bundle = load_export_bundle_sync(session, org_id, project_id)
    safe_name = "".join(ch for ch in bundle.project_name if ch.isalnum() or ch in (" ", "-", "_")).strip() or "project"
    if export_format == "xlsx":
        data = build_xlsx_bytes(bundle)
        return data, f"{safe_name}-quantities.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if export_format == "pdf":
        data = build_pdf_bytes(bundle)
        return data, f"{safe_name}-quantities.pdf", "application/pdf"
    raise ValueError(f"unsupported format: {export_format}")
